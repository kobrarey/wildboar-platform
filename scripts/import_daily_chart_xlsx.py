from __future__ import annotations

import argparse
from datetime import date, datetime, time, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.db import SessionLocal
from app.models import Fund, FundChartDaily


SHEET_NAME = "Лист1"

BLOCKS = [
    {
        "fund_code": "defi_sniper",
        "label": "WB DeFi Sniper",
        "start_col": 1,   # A:E
    },
    {
        "fund_code": "btc_fund",
        "label": "WB BTC",
        "start_col": 7,   # G:K
    },
    {
        "fund_code": "wb10",
        "label": "WB 10",
        "start_col": 13,  # M:Q
    },
    {
        "fund_code": "wb_test",
        "label": "WB Test",
        "start_col": 19,  # S:W
    },
]

DATA_START_ROW = 3


def _to_decimal(value: Any, *, row_idx: int, field_name: str, fund_code: str) -> Decimal:
    if value is None or value == "":
        raise ValueError(f"{fund_code} row={row_idx}: missing {field_name}")

    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(
            f"{fund_code} row={row_idx}: invalid {field_name}={value!r}"
        ) from exc


def _parse_excel_date(value: Any, *, row_idx: int, fund_code: str) -> datetime:
    if value is None or value == "":
        raise ValueError(f"{fund_code} row={row_idx}: missing date")

    if isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, time.min)
    elif isinstance(value, (int, float)):
        parsed = from_excel(value)
        if isinstance(parsed, datetime):
            dt = parsed
        elif isinstance(parsed, date):
            dt = datetime.combine(parsed, time.min)
        else:
            raise ValueError(f"{fund_code} row={row_idx}: unsupported Excel date={value!r}")
    else:
        s = str(value).strip()
        if not s:
            raise ValueError(f"{fund_code} row={row_idx}: missing date")

        if s.endswith("Z"):
            s = s[:-1] + "+00:00"

        try:
            dt = datetime.fromisoformat(s)
        except ValueError as exc:
            raise ValueError(f"{fund_code} row={row_idx}: invalid date={value!r}") from exc

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    return dt.astimezone(timezone.utc).replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0,
    )


def _get_funds_map() -> dict[str, int]:
    codes = [b["fund_code"] for b in BLOCKS]

    with SessionLocal() as db:
        rows = db.query(Fund).filter(Fund.code.in_(codes)).all()

    out = {row.code: int(row.id) for row in rows}

    missing = [code for code in codes if code not in out]
    if missing:
        raise RuntimeError("Funds not found in DB: " + ", ".join(missing))

    return out


def _read_block_rows(ws, *, fund_code: str, start_col: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []

    for row_idx in range(DATA_START_ROW, ws.max_row + 1):
        raw_date = ws.cell(row=row_idx, column=start_col).value
        raw_open = ws.cell(row=row_idx, column=start_col + 1).value
        raw_high = ws.cell(row=row_idx, column=start_col + 2).value
        raw_low = ws.cell(row=row_idx, column=start_col + 3).value
        raw_close = ws.cell(row=row_idx, column=start_col + 4).value

        # Empty line inside one block means no more data for that block.
        if (
            raw_date is None
            and raw_open is None
            and raw_high is None
            and raw_low is None
            and raw_close is None
        ):
            continue

        ts_utc = _parse_excel_date(raw_date, row_idx=row_idx, fund_code=fund_code)

        rows.append(
            {
                "fund_code": fund_code,
                "ts_utc": ts_utc,
                "open": _to_decimal(raw_open, row_idx=row_idx, field_name="open", fund_code=fund_code),
                "high": _to_decimal(raw_high, row_idx=row_idx, field_name="high", fund_code=fund_code),
                "low": _to_decimal(raw_low, row_idx=row_idx, field_name="low", fund_code=fund_code),
                "close": _to_decimal(raw_close, row_idx=row_idx, field_name="close", fund_code=fund_code),
                "volume": None,
            }
        )

    return rows


def import_daily_chart_xlsx(*, file_path: Path) -> list[dict[str, Any]]:
    if not file_path.exists():
        raise FileNotFoundError(f"XLSX file not found: {file_path}")

    wb = load_workbook(file_path, data_only=True, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {SHEET_NAME}")

    ws = wb[SHEET_NAME]
    funds_map = _get_funds_map()

    summaries: list[dict[str, Any]] = []

    with SessionLocal() as db:
        for block in BLOCKS:
            fund_code = block["fund_code"]
            fund_id = funds_map[fund_code]

            rows = _read_block_rows(
                ws,
                fund_code=fund_code,
                start_col=block["start_col"],
            )

            rows_written = 0

            for row in rows:
                item = {
                    "fund_id": fund_id,
                    "ts_utc": row["ts_utc"],
                    "open": row["open"],
                    "high": row["high"],
                    "low": row["low"],
                    "close": row["close"],
                    "volume": None,
                }

                stmt = (
                    pg_insert(FundChartDaily.__table__)
                    .values(**item)
                    .on_conflict_do_update(
                        index_elements=["fund_id", "ts_utc"],
                        set_={
                            "open": item["open"],
                            "high": item["high"],
                            "low": item["low"],
                            "close": item["close"],
                            "volume": None,
                        },
                    )
                )

                db.execute(stmt)
                rows_written += 1

            dates = [row["ts_utc"] for row in rows]

            summaries.append(
                {
                    "fund_code": fund_code,
                    "rows_read": len(rows),
                    "rows_inserted_or_updated": rows_written,
                    "min_date": min(dates).date().isoformat() if dates else None,
                    "max_date": max(dates).date().isoformat() if dates else None,
                }
            )

        db.commit()

    return summaries


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import daily fund chart candles from XLSX.")
    parser.add_argument(
        "--file",
        required=True,
        help="Path to XLSX file with daily candles.",
    )
    return parser.parse_args()


def main() -> int:
    load_dotenv()

    args = parse_args()
    summaries = import_daily_chart_xlsx(file_path=Path(args.file))

    print("Daily chart XLSX import summary:")
    for summary in summaries:
        print(
            "fund_code={fund_code} rows_read={rows_read} "
            "rows_inserted_or_updated={rows_inserted_or_updated} "
            "min_date={min_date} max_date={max_date}".format(**summary)
        )

    return 0


if __name__ == "__main__":
    raise SystemExit(main())